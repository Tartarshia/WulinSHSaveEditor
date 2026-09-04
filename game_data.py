"""Read the installed game's official item table; no table is bundled or uploaded."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re
from zipfile import ZipFile

from openpyxl import load_workbook


def item_category(item_type: str) -> str:
    """Map the game's detailed ItemType enum to the familiar inventory groups."""
    if item_type.startswith("Equip_Weapon_Pet") or item_type.startswith("Equip_Armor_Pet") or item_type.startswith("Equip_Amulet_Pet"):
        return "宠物装备"
    if item_type.startswith("Equip_Weapon"):
        return "武器"
    if item_type.startswith("Equip_Armor"):
        return "防具"
    if item_type.startswith("Equip_Amulet"):
        return "饰品"
    if item_type.startswith("KungfuBook"):
        return "武功秘籍"
    if item_type.startswith("Consumeable_Material"):
        return "材料"
    if item_type.startswith("Consumeable_Recipe"):
        return "配方"
    if item_type.startswith("Consumeable"):
        return "消耗品"
    if item_type.startswith("Misc_HidWep"):
        return "暗器"
    if item_type.startswith("Misc_Map"):
        return "地图"
    if item_type.startswith("Misc_Poison"):
        return "毒物"
    if item_type.startswith("Misc_Quest"):
        return "任务物品"
    if item_type.startswith("Misc_Treasure"):
        return "珍宝"
    return "杂项" if item_type != "None" else "未分类"


def load_item_catalog(game_root: Path) -> dict[int, tuple[str, str]]:
    archive = game_root / "Wulin_Data/StreamingAssets/ModEditor/ModEditor.zip"
    if not archive.is_file():
        raise FileNotFoundError(f"未找到游戏自带物品表：{archive}")
    with ZipFile(archive) as bundle:
        candidates = [info for info in bundle.infolist() if info.filename.lower().endswith(".xlsx") and "/xls/1/" in info.filename.replace("\\", "/")]
        for info in candidates:
            workbook = load_workbook(BytesIO(bundle.read(info)), read_only=True, data_only=True)
            if "ItemData" not in workbook.sheetnames:
                workbook.close()
                continue
            sheet = workbook["ItemData"]
            # The shipped ItemData table defines uid in B and display name in C.
            result: dict[int, tuple[str, str]] = {}
            for row in sheet.iter_rows(min_row=4, min_col=2, max_col=10, values_only=True):
                item_id, name, item_type = row[0], row[1], row[8]
                try:
                    item_id = int(item_id)
                except (TypeError, ValueError):
                    continue
                if isinstance(name, str) and name:
                    result[item_id] = (name, item_category(str(item_type or "None")))
            workbook.close()
            if result:
                return result
    raise RuntimeError("游戏 ModEditor 中没有找到 ItemData 物品表")


def load_item_names(game_root: Path) -> dict[int, str]:
    """Compatibility helper for callers which only need display names."""
    return {item_id: value[0] for item_id, value in load_item_catalog(game_root).items()}


def load_character_names(game_root: Path) -> dict[int, str]:
    """Return template ID -> display name from the installed CharacterPoolData."""
    archive = game_root / "Wulin_Data/StreamingAssets/ModEditor/ModEditor.zip"
    if not archive.is_file():
        raise FileNotFoundError(f"未找到游戏自带角色表：{archive}")
    with ZipFile(archive) as bundle:
        candidates = [info for info in bundle.infolist() if info.filename.lower().endswith(".xlsx") and "/xls/1/" in info.filename.replace("\\", "/")]
        for info in candidates:
            workbook = load_workbook(BytesIO(bundle.read(info)), read_only=True, data_only=True)
            if "CharacterPoolData" not in workbook.sheetnames:
                workbook.close()
                continue
            sheet = workbook["CharacterPoolData"]
            result: dict[int, str] = {}
            for template_id, name in sheet.iter_rows(min_row=4, min_col=2, max_col=3, values_only=True):
                try:
                    template_id = int(template_id)
                except (TypeError, ValueError):
                    continue
                if isinstance(name, str) and name:
                    result[template_id] = name
            workbook.close()
            if result:
                return result
    raise RuntimeError("游戏 ModEditor 中没有找到 CharacterPoolData 角色表")


def load_faction_names(game_root: Path) -> dict[int, str]:
    archive = game_root / "Wulin_Data/StreamingAssets/ModEditor/ModEditor.zip"
    if not archive.is_file():
        raise FileNotFoundError(f"未找到游戏自带势力表：{archive}")
    with ZipFile(archive) as bundle:
        for info in bundle.infolist():
            if not info.filename.lower().endswith(".xlsx"):
                continue
            workbook = load_workbook(BytesIO(bundle.read(info)), read_only=True, data_only=True)
            if "FactionData" not in workbook.sheetnames:
                workbook.close()
                continue
            result: dict[int, str] = {}
            for faction_id, name in workbook["FactionData"].iter_rows(min_row=4, min_col=2, max_col=3, values_only=True):
                try:
                    faction_id = int(faction_id)
                except (TypeError, ValueError):
                    continue
                if isinstance(name, str) and name:
                    result[faction_id] = name
            workbook.close()
            if result:
                return result
    raise RuntimeError("游戏 ModEditor 中没有找到 FactionData 势力表")


def load_job_factions(game_root: Path) -> dict[int, str]:
    """Return faction job ID -> parent faction display name."""
    archive = game_root / "Wulin_Data/StreamingAssets/ModEditor/ModEditor.zip"
    with ZipFile(archive) as bundle:
        for info in bundle.infolist():
            if not info.filename.lower().endswith(".xlsx"):
                continue
            workbook = load_workbook(BytesIO(bundle.read(info)), read_only=True, data_only=True)
            if "FactionJobData" not in workbook.sheetnames:
                workbook.close()
                continue
            result: dict[int, str] = {}
            for job_id, _name, faction_name in workbook["FactionJobData"].iter_rows(min_row=4, min_col=2, max_col=4, values_only=True):
                try:
                    job_id = int(job_id)
                except (TypeError, ValueError):
                    continue
                if isinstance(faction_name, str) and faction_name:
                    result[job_id] = faction_name
            workbook.close()
            if result:
                return result
    raise RuntimeError("游戏 ModEditor 中没有找到 FactionJobData 势力职位表")


def load_kungfu_catalog(game_root: Path) -> dict[int, tuple[str, int]]:
    """Return kungfu ID -> (display name, maximum level)."""
    archive = game_root / "Wulin_Data/StreamingAssets/ModEditor/ModEditor.zip"
    with ZipFile(archive) as bundle:
        for info in bundle.infolist():
            if not info.filename.lower().endswith(".xlsx"):
                continue
            workbook = load_workbook(BytesIO(bundle.read(info)), read_only=True, data_only=True)
            if "KungfuData" not in workbook.sheetnames:
                workbook.close()
                continue
            result: dict[int, tuple[str, int]] = {}
            for row in workbook["KungfuData"].iter_rows(min_row=4, min_col=2, max_col=9, values_only=True):
                try:
                    kungfu_id, max_level = int(row[0]), int(row[7])
                except (TypeError, ValueError):
                    continue
                if isinstance(row[1], str) and row[1]:
                    result[kungfu_id] = (row[1], max_level)
            workbook.close()
            if result:
                return result
    raise RuntimeError("游戏 ModEditor 中没有找到 KungfuData 武功表")


def load_life_ability_catalog(game_root: Path) -> dict[str, int]:
    """Return player life-ability property key -> official maximum level."""
    archive = game_root / "Wulin_Data/StreamingAssets/ModEditor/ModEditor.zip"
    with ZipFile(archive) as bundle:
        for info in bundle.infolist():
            if not info.filename.lower().endswith(".xlsx"):
                continue
            workbook = load_workbook(BytesIO(bundle.read(info)), read_only=True, data_only=True)
            if "GameEnvConstData" not in workbook.sheetnames:
                workbook.close()
                continue
            result: dict[str, int] = {}
            for row in workbook["GameEnvConstData"].iter_rows(values_only=True):
                if len(row) < 4 or not isinstance(row[2], str) or not row[2].startswith("能力_") or not row[2].endswith("Max"):
                    continue
                ability_key = row[2][:-3]
                if ability_key.count("_") != 2:
                    continue
                try:
                    result[ability_key] = int(row[3])
                except (TypeError, ValueError):
                    continue
            workbook.close()
            if result:
                return result
    raise RuntimeError("游戏 ModEditor 中没有找到主角生活能力上限表")


def load_jianghu_experience_max(game_root: Path) -> int:
    """Read the Jianghu Knowledge pool cap from the installed game table."""
    archive = game_root / "Wulin_Data/StreamingAssets/ModEditor/ModEditor.zip"
    with ZipFile(archive) as bundle:
        for info in bundle.infolist():
            if not info.filename.lower().endswith(".xlsx"):
                continue
            workbook = load_workbook(BytesIO(bundle.read(info)), read_only=True, data_only=True)
            if "GameEnvConstData" not in workbook.sheetnames:
                workbook.close()
                continue
            for row in workbook["GameEnvConstData"].iter_rows(values_only=True):
                if len(row) >= 4 and row[2] == "江湖历练AddMax":
                    try:
                        maximum = int(row[3])
                    except (TypeError, ValueError):
                        break
                    workbook.close()
                    return maximum
            workbook.close()
    raise RuntimeError("游戏 ModEditor 中没有找到江湖历练上限")


def load_recipe_catalog(game_root: Path) -> dict[int, str]:
    """Return active recipe ID -> name, excluding rows marked skip by the game data."""
    archive = game_root / "Wulin_Data/StreamingAssets/ModEditor/ModEditor.zip"
    with ZipFile(archive) as bundle:
        for info in bundle.infolist():
            if not info.filename.lower().endswith(".xlsx"):
                continue
            workbook = load_workbook(BytesIO(bundle.read(info)), read_only=True, data_only=True)
            if "NewRecipeData" not in workbook.sheetnames:
                workbook.close()
                continue
            result: dict[int, str] = {}
            for remark, recipe_id, name in workbook["NewRecipeData"].iter_rows(min_row=4, min_col=1, max_col=3, values_only=True):
                if remark == "skip":
                    continue
                try:
                    recipe_id = int(recipe_id)
                except (TypeError, ValueError):
                    continue
                if isinstance(name, str) and name:
                    result[recipe_id] = name
            workbook.close()
            if result:
                return result
    raise RuntimeError("游戏 ModEditor 中没有找到有效配方表")


def load_team_stamina_bonuses(game_root: Path) -> dict[str, float]:
    """Return Buff data ID -> additive team-stamina-cap bonus from game data."""
    archive = game_root / "Wulin_Data/StreamingAssets/ModEditor/ModEditor.zip"
    pattern = re.compile(r"队伍体力Max@团队\|(-?\d+(?:\.\d+)?)")
    with ZipFile(archive) as bundle:
        for info in bundle.infolist():
            if not info.filename.lower().endswith(".xlsx"):
                continue
            workbook = load_workbook(BytesIO(bundle.read(info)), read_only=True, data_only=True)
            if "BuffData" not in workbook.sheetnames:
                workbook.close()
                continue
            result: dict[str, float] = {}
            for row in workbook["BuffData"].iter_rows(min_row=4, values_only=True):
                if len(row) < 3 or not isinstance(row[2], str):
                    continue
                bonus = sum(float(match.group(1)) for value in row if isinstance(value, str) for match in pattern.finditer(value))
                if bonus:
                    result[row[2]] = bonus
            workbook.close()
            if result:
                return result
    raise RuntimeError("游戏 ModEditor 中没有找到队伍体力上限 Buff 数据")
